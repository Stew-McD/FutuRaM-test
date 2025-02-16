#%%
import os
import pandas as pd
import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font


#%% 
# Create a new workbook
wb = xl.Workbook()

# Create a new worksheet and rename it to "Information"
info = wb.create_sheet("Information", index=0)
wb.remove(wb["Sheet"])

info.cell(row=2, column=1).value = "This file contains the templates for the different matter classes. Please fill in the information for each matter in the corresponding sheet."
info.cell(row=2, column=1).font = Font(size=18, bold=True)
info.row_dimensions[2].height = 30
info.cell(row=3, column=1).value = "The following matter classes are available:"
info.cell(row=4, column=1).value = "1. Elements"
info.cell(row=5, column=1).value = "2. Compounds"
info.cell(row=6, column=1).value = "3. Materials"
info.cell(row=7, column=1).value = "4. Components"
info.cell(row=8, column=1).value = "5. Products"

info.cell(row=12, column=1).value = "Each matter should have it's own datasheet"
info.cell(row=12, column=1).font = Font(size=14, bold=True)
info.row_dimensions[12].height = 30
info.cell(row=13, column=1).value = '=HYPERLINK("../datatemplates/matterDatasheetTemplate.xlsx", "matter datasheet template")'
drawing = Image("OverviewOfClassesInTheModel.png")
info.add_image(drawing, "B9")

# Define the standard columns for all sheets
columns = ['name', 'code', "matter_type", 'criticality', "WasteStreams", 'code_EC','code_UNSPC', 'code_CAS', "prod_category", "prod_sub_category", "ISIC_name", "ISIC", "CPC_name", "CPC", "code_EI", 'tags', 'notes' ,'datasheet']

# Add extra columns for each subclass if needed
columns_ele = columns + []
columns_com = columns + []
columns_mat = columns + []
columns_cpt = columns + []
columns_pro = columns + []

# Dictionary for making the sheets for each subclass
subclasses = {
    "matters_template": columns,
    "elements": columns_ele,
    "compounds": columns_com,
    "materials": columns_mat,
    "components": columns_cpt,
    "products": columns_pro
}

# Make the sheets 
for subclass, cols in subclasses.items():
    sheet = wb.create_sheet(subclass)
    for i, col in enumerate(cols):
        # Set the value of the cell
        sheet.cell(row=1, column=i+1).value = col
        # Set the font style for the header row
        sheet.cell(row=1, column=i+1).font = xl.styles.Font(bold=True)
        # Set the column width
        sheet.column_dimensions[xl.utils.get_column_letter(i+1)].width = 15
        # Set the background color for the header row
        sheet.cell(row=1, column=i+1).fill = xl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
        # Set the autosort for the header row
        sheet.auto_filter.ref = sheet.dimensions

    # Add an example row with some data
    example_row = ['Example', '123', 'material', 'high', 'BATT, ELV', '123456', '123456', '123456-78-9', 'Metals', 'subcategory1', 'ISIC_name1', 'ISIC1', 'CPC_name1', 'CPC1', '123', 'tag1, tag2', 'notes', '=HYPERLINK("datasheets/"&B2".xlsx"']
    for i, col in enumerate(example_row):
        sheet.cell(row=2, column=i+1).value = col
        sheet.cell(row=2, column=i+1).fill = xl.styles.PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')

# Save the file
if not os.path.exists("datatemplates"):
    os.makedirs("datatemplates")
wb.save("data/ListOfAllmatters.xlsx")
wb.close()