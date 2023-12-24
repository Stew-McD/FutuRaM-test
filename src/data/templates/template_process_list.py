#%%
import os
import pandas as pd
import openpyxl as xl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font



#%% 
# Create a new workbook

def make_workbook_process_list():

    if not os.path.exists("../data/templates"):
        os.makedirs("../data/templates")
    
    if not os.isfile("../data/templates/ProcessDatasheetTemplate.xlsx"):
        wb = xl.Workbook()

        # Create a new worksheet and rename it to "Information"
        info = wb.create_sheet("Information", index=0)
        wb.remove(wb["Sheet"])

        info.cell(row=1, column=1).value = "This file contains the templates for process class. Please fill in the information for each process \n each process must also have a datasheet with complete information about its flows and parameters"
        info.cell(row=1, column=1).font = Font(size=18, bold=True)
        info.row_dimensions[1].height = 40
        info.cell(row=3, column=1).value = "The processes should contain the following information:"

        info.cell(row=4, column=1).value = "1. Name"
        info.cell(row=5, column=1).value = "2. Code"
        info.cell(row=6, column=1).value = "3. Category"
        info.cell(row=7, column=1).value = "4. Subcategory"
        info.cell(row=8, column=1).value = "5. Tags"
        info.cell(row=9, column=1).value = "6. Notes"
        info.cell(row=10, column=1).value = "7. Datasheet link"
        info.cell(row=12, column=1).value = "The template for the process datasheet is in a separate file in the templates folder"
        info.cell(row=12, column=1).font = Font(size=14, bold=True)
        info.row_dimensions[12].height = 30
        info.cell(row=13, column=1).value = '=HYPERLINK("../datatemplates/ProcessDatasheetTemplate.xlsx", "Process datasheet template")'

        drawing = Image("OverviewOfClassesInTheModel.png")
        info.add_image(drawing, "A14")

        # Define the standard columns for all sheets
        columns = ['name', 'code', "category", 'subcategory', 'WasteStreams' 'tags', 'notes' ,'datasheet']

        # Add extra columns for each subclass if needed
        columns_process = columns + []
        # Dictionary for making the sheets for each subclass
        subclasses = {
            "processes": columns,
        }

        # Make the sheets 
        for subclass, cols in subclasses.items():
            sheet = wb.create_sheet(subclass)
            for i, col in enumerate(cols):
                # Set the value of the cell
                sheet.cell(row=1, column=i+1).value = col
                # Set the font style for the header row
                sheet.cell(row=1, column=i+1).font = xl.styles.Font(bold=True)
                # Set the column width
                sheet.column_dimensions[xl.utils.get_column_letter(i+1)].width = 15
                # Set the background color for the header row
                sheet.cell(row=1, column=i+1).fill = xl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
                # Set the autosort for the header row
                sheet.auto_filter.ref = sheet.dimensions

            # Add an example row with some data
            example_row = ['Process1', 'P1', 'Category1', 'Subcategory1', 'tag1, tag2, tag3', 'This is a note', '=HYPERLINK("../data/process/$B2.xlsx";"datasheet")']
            for i, col in enumerate(example_row):
                sheet.cell(row=2, column=i+1).value = col
                sheet.cell(row=2, column=i+1).fill = xl.styles.PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')


        # Save the file
        wb.save("../data/templates/ListOfAllProcesses.xlsx")
        wb.close()