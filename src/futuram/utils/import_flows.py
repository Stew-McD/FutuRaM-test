"""
import_flows.py

This module contains functions for importing flows to a model from a file.

Functions:
- import_flows_xlsx(filename, model):  Import flows from an xlsx file.

Dependencies:
- openpyxl
- ../classes/model.py
- ../classes/processes.py
- ../classes/matter.py
- ../classes/flows.py

"""


import openpyxl

from ..classes.flows import Flow
from ..classes.processes import Process


def import_flows_xlsx(filename, model):
    """
    Import flows from an xlsx file.
    The xlsx file should have the following structure:
    process_from,process_to,composition

    Parameters
    ----------
    filename : str
        The path to the xlsx file.

    model : Model
        The model object to add the flows to.
    
    """
    print(
        f'\n\n{"-" * 70}\n\t Importing flows to model "{model.name}" from {filename}\n{"-" * 70}'
    )

    # set the path to the file
    path = filename
    workbook = openpyxl.load_workbook(path, data_only=True)
    # Select the worksheet
    worksheet = workbook.worksheets[0]
    # Get the headers
    headers = [cell.value for cell in worksheet[1]]
    # Convert each row (tc) to a dictionary
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if worksheet.cell(row=1, column=i + 1).data_type == "s":
                row_data[headers[i]] = str(value)
            elif worksheet.cell(row=1, column=i + 1).data_type == "n":
                row_data[headers[i]] = float(value)
            elif worksheet.cell(row=1, column=i + 1).data_type == "d":
                row_data[headers[i]] = value.date()
            elif worksheet.cell(row=1, column=i + 1).data_type == "b":
                row_data[headers[i]] = bool(value)
            elif worksheet.cell(row=1, column=i + 1).data_type == "f":
                row_data[headers[i]] = worksheet.cell(
                    row=row[0].row, column=row[0].column
                ).value
            else:
                row_data[headers[i]] = None
        data.append(row_data)

    process_dict = {process.name: process for process in model.processes.values()}

    count = 0
    for flow in data:
        count += 1
        print(
            f'Importing flow {count}/{len(data)}: {flow["process_from"]} -> {flow["process_to"]}'
        )
        new_flow = Flow(
            model, flow["process_from"], flow["process_to"], flow["composition"]
        )

        if flow["process_from"] in process_dict:
            process = process_dict[flow["process_from"]]
            process.add_output(new_flow)

        else:
            print("Process not found, adding: " + flow["process_from"])
            process = Process(flow["process_from"])
            process.add_output(new_flow)
            model.add_process(process)
            # process.add_to_model(model)

        if flow["process_to"] in process_dict:
            process = process_dict[flow["process_to"]]
            process.add_input(new_flow)
        else:
            print("Process not found, adding: " + flow["process_to"])
            process = Process(flow["process_from"])
            process.add_input(new_flow)
            model.add_process(process)
            process.add_to_model(model)

        model.get_flows()
