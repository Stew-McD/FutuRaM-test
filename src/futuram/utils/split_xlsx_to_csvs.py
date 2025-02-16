import openpyxl
import csv
import os

def xlsx_to_csvs(filename):
    """
    This function splits an xlsx file into separate CSV files for each sheet.

    Parameters
    ----------
    filename : str
        The path to the xlsx file.
    
    Returns
    -------
    folder_name : str
        The name of the folder where the CSV files were created.


    """
    print(f'\n\n{"*" * 60}\n{"*" * 4} Splitting {os.path.basename(filename)} into CSV files {"*" * 4}\n{"*" * 60}')
    # set the path to the file
    path = filename
    workbook = openpyxl.load_workbook(path, data_only=True)
    # Get the sheet names
    sheet_names = workbook.sheetnames
    # Create a new directory for the CSV files
    folder_name = f'{os.path.splitext(filename)[0]}-split'
    os.makedirs(folder_name, exist_ok=True)
    # Loop through each sheet and create a CSV file
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        with open(f'{folder_name}/{sheet_name}.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            # Write the headers to the CSV file
            headers = [cell.value for cell in worksheet[1]]
            writer.writerow(headers)
            # Write the data to the CSV file
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row[0] or 'total' not in row[0].lower():
                    if row[-1] is None:
                        row = row[:-1]
                    writer.writerow(row)
        # Find the last non-empty row
        last_row = worksheet.max_row
        while last_row > 1 and all(cell.value is None for cell in worksheet[last_row]):
            last_row -= 1
        # Remove empty rows at the end of the CSV file
        with open(f'{folder_name}/{sheet_name}.csv', 'r+', newline='') as csvfile:
            data = csvfile.read().splitlines(True)
            csvfile.seek(0)
            csvfile.writelines(data[:last_row])
            csvfile.truncate()
    print(f'\nFiles were created in {folder_name}')
    print('\nSheets extracted:')
    for sheet_name in sheet_names: print(f'\t\t{sheet_name}')

    return folder_name