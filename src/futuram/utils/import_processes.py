'''
import_processes.py

This module contains functions to import processes from
xlsx files to a model.

Methods:
    import_processes_xlsx(filename, model)

Dependencies:
    openpyxl

#! there is probably no reason to be dealing with xlsx files directly in the model, but this is likely how the data will arrive. We could also convert the xlsx files to csvs and import them that way.

:noindex:
'''

import openpyxl
from ..classes.processes import Process

def import_processes_xlsx(filename, model):
    """
    Import a xlsx file containing process data and create process objects for each row

    Parameters
    ----------

    filename (str): The path to the xlsx file.
    model (Model): The model object to add the processes to.


    """
    print(f'\n\n{"-" * 60}\n   Importing processes to model \"{model.name}\" from {filename}\n{"-" * 60}\n')


    # set the path to the file
    path = filename
    workbook = openpyxl.load_workbook(path, data_only=True)
    # Select the worksheet
    worksheet = workbook.worksheets[0]
    # Get the headers
    headers = [cell.value for cell in worksheet[1]]
    # Convert each row (process) to a dictionary
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if worksheet.cell(row=1, column=i+1).data_type == 's':
                row_data[headers[i]] = str(value)
            elif worksheet.cell(row=1, column=i+1).data_type == 'n':
                row_data[headers[i]] = float(value)
            elif worksheet.cell(row=1, column=i+1).data_type == 'd':
                row_data[headers[i]] = value.date()
            elif worksheet.cell(row=1, column=i+1).data_type == 'b':
                row_data[headers[i]] = bool(value)
            elif worksheet.cell(row=1, column=i+1).data_type == 'f':
                row_data[headers[i]] = worksheet.cell(row=row[0].row, column=row[0].column)._value
            else:
                row_data[headers[i]] = None
        data.append(row_data)

    print(f"** Importing {len(data)} processes **\n")
    for process in data:
        new_process = Process(process['name'])
        new_process.uuid = process['uuid']
        new_process.description = process['description']
        new_process.tags = process['tags']
        new_process.WS = process['WS']
        new_process.transformation_level = process['transformation_level']
        new_process.consumption_energy = process['consumption_energy']
        new_process.consumption_water = process['consumption_water']
        new_process.cost_operation = process['cost_operation']

        new_process.add_to_model(model)
        print(f"\t{new_process.name}")
