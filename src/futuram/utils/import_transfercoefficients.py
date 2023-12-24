"""
import_transfercoefficients.py

This module contains functions to import transfer coefficients from
xlsx files to a model.

Methods:
    import_transfercoefficients_xlsx(filename, model)

Dependencies:
    openpyxl
"""

import openpyxl

#! as with the other import functions, we should probably be doing this with csvs instead of xlsx files.


def import_transfercoefficients_xlsx(filename, model):
    """
    Import a xlsx file containing transfer coefficient data and add them to the model
    Args:
        filename (str): The path to the xlsx file.
        model (Model): The model object to add the transfer coefficients to.

    The xlsx file should have the following columns:
        process (str): The name of the process.
        flow_input (str): The name of the input flow.
        flow_output (str): The name of the output flow.
        transfer_coefficient (float): The transfer coefficient.
        uncertainty (float): The uncertainty of the transfer coefficient.

    """
    print(
        f'\n\n{"*" * 90}\n  Importing transfer coefficients to model "{model.name}" from {filename}\n{"*" * 90}'
    )

    # set the path to the file
    path = filename
    workbook = openpyxl.load_workbook(path, data_only=True)
    # Select the worksheet
    worksheet = workbook.worksheets[0]
    # Get the headers
    headers = [cell.value for cell in worksheet[1]]
    # Convert each row (tc) to a dictionary
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if worksheet.cell(row=1, column=i + 1).data_type == "s":
                row_data[headers[i]] = str(value)
            elif worksheet.cell(row=1, column=i + 1).data_type == "n":
                row_data[headers[i]] = float(value)
            elif worksheet.cell(row=1, column=i + 1).data_type == "d":
                row_data[headers[i]] = value.date()
            elif worksheet.cell(row=1, column=i + 1).data_type == "b":
                row_data[headers[i]] = bool(value)
            elif worksheet.cell(row=1, column=i + 1).data_type == "f":
                row_data[headers[i]] = worksheet.cell(
                    row=row[0].row, column=row[0].column
                )._value
            else:
                row_data[headers[i]] = None
        data.append(row_data)

    process_dict = {process.name: process for process in model.processes.values()}

    for tc in data:
        process_name = tc["process"]
        if process_name in model.processes:
            process = model.processes[process_name]
            process.add_transfer_coefficient(
                tc["flow_input"],
                tc["flow_output"],
                float(tc["transfer_coefficient"]),
                float(tc["uncertainty"]),
            )
            print(
                f"\t* {process.name} \t\t({tc['flow_input']} --{round(float(tc['transfer_coefficient']),2)}--> {tc['flow_output']})"
            )
        else:
            print(f" ****** Process not found: {process_name} ******")

    # return data
